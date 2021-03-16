"""
Excel Write
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Font

class Excel:
    def __init__(self) -> None:
        self.file_name = ""

    def __init__(self, file_name="") -> None:
        self.file_name = file_name

    def write(self, sheet_name="Sheet1", data_array = [], write_mode='a'):
        if(not os.path.exists(self.file_name)) or (write_mode == 'o'):
            print("[+] Info: Create {}".format(self.file_name))
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            worksheet.append(data_array)
            workbook.save(self.file_name)
        
        else:
            try:
                workbook  = openpyxl.load_workbook(self.file_name)
                worksheet = workbook[sheet_name]
                worksheet.append(data_array)
                workbook.save(self.file_name)
            except:
                print("[-] Error: Write item to {} failed!".format(self.file_name))
                return

    def set_col_width(self, sheet_name="Sheet1", cols_width={'A':20}):

        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))
        

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        for key in cols_width.keys():
            width = cols_width[key] if(type(cols_width[key]) == int) else 10
            worksheet.column_dimensions[key].width = width

        workbook.save(self.file_name)

    def set_row_height(self, sheet_name="Sheet1", rows_height={1:20}):

        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))
        

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        for key in rows_height.keys():
            height = rows_height[key] if(type(rows_height[key]) == int) else 10
            worksheet.row_dimensions[key].height = height

        workbook.save(self.file_name)

    def set_center_row(self, sheet_name="Sheet1", row_num=1):
        
        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        if(row_num > worksheet.max_row):
            print("[-] Warning: {} bigger than max row in {} of file {}!".format(row_num, sheet_name, self.file_name))
            return

        for col_i in range(worksheet.max_column):
            worksheet.cell(row=row_num, column=col_i+1).alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(self.file_name)

    def set_center_col(self, sheet_name="Sheet1", col_num=1):
        
        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        if(col_num > worksheet.max_column):
            print("[-] Warning: {} bigger than max row in {} of file {}!".format(col_num, sheet_name, self.file_name))
            return

        for row_i in range(worksheet.max_row):
            worksheet.cell(row=row_i+1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(self.file_name)

    def set_font(self, font_style, sheet_name="Sheet1", row_arr=[], col_arr=[], all=False):
        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        if all:
            row_arr = range(1, worksheet.max_row+1)
            col_arr = range(1, worksheet.max_column+1)

        for row in row_arr:
            if row > worksheet.max_row:
                break
            for col in col_arr:
                if col > worksheet.max_column:
                    break
                worksheet.cell(row=row, column=col).font = font_style

        workbook.save(self.file_name)


    def set_col_auto_wraptext_center(self, sheet_name="Sheet1", col_nums=[1], all=False):
        
        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        if all:
            col_nums = range(1, worksheet.max_column+1)

        for col_num in col_nums:
            if col_num > worksheet.max_column:
                break
            for row_i in range(worksheet.max_row):
                worksheet.cell(row=row_i+1, column=col_num).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        workbook.save(self.file_name)

    def set_row_auto_wraptext_center(self, sheet_name="Sheet1", row_nums=[1], all=False):
        
        if not os.path.exists(self.file_name):
            print("[-] Error: {} does not exists!".format(self.file_name))

        workbook  = openpyxl.load_workbook(self.file_name)
        worksheet = workbook[sheet_name]

        if all:
            row_nums = range(1, worksheet.max_row+1)

        for row_num in row_nums:
            if row_num > worksheet.max_row:
                break
            for col_i in range(worksheet.max_column):
                worksheet.cell(row=row_num, column=col_i+1).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        workbook.save(self.file_name)

def main():
    file_name = "test.xlsx"
    font      = Font(name='Times New Roman',
                     size=11,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='00000000')

    excel0    = Excel(file_name=file_name)
    excel0.write(data_array=["namenamenamenamenamenamenamenamenamenamenamenamenamenamename", "idnamenamenamenamename"])
    # excel0.set_col_width(cols_width={'A':10})
    excel0.set_col_auto_wraptext_center(col_nums=[1,2])
    excel0.set_row_auto_wraptext_center(row_nums=[1,2])
    # excel0.set_row_height(rows_height={1:20})
    # excel0.set_center_row(row_num=1)
    # excel0.set_center_col(col_num=1)
    excel0.set_font(font, all=True)


if __name__ == "__main__":
    main()
